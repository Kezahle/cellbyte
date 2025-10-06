from pathlib import Path
import pandas as pd

from config.config import settings

class Exporter:
    """Handles artifact format conversions and exports."""

    def convert_format(self, source_path: Path, target_format: str) -> Path:
        if not source_path.exists():
            raise FileNotFoundError(f"Source file not found: {source_path}")

        source_format = source_path.suffix.lstrip('.')
        
        # Validate conversion is supported
        supported_conversions = {
            ('csv', 'xlsx'): self._csv_to_excel,
        }
        
        if (source_format, target_format) not in supported_conversions:
            raise ValueError(f"Conversion from '{source_format}' to '{target_format}' is not supported. Supported: CSV→XLSX")
        
        output_path = source_path.with_suffix(f".{target_format}")
        converter = supported_conversions[(source_format, target_format)]
        return converter(source_path, output_path)

    def _csv_to_excel(self, source_path: Path, output_path: Path) -> Path:
        """Converts a CSV file to a formatted Excel file."""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise ImportError("Please install 'openpyxl' to export to Excel.")

        df = pd.read_csv(source_path)
        df.to_excel(output_path, index=False, sheet_name='Data')

        wb = load_workbook(output_path)
        ws = wb.active
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except: pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        return output_path